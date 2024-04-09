import sys
from PyQt5.QtWidgets import *
from PyQt5 import uic
import openpyxl as op
from pathlib import Path
import shutil
import pyperclip
from utils.backend import BackendManager
from API.building import *
from template.buildings import *
import os
import ast
import UI.我不知道我抄csdn的_rc
import tempfile
from PyQt5.QtCore import Qt
import qfluentwidgets

# 这是一个示例 Python 脚本。
# 按 Shift+F10 执行或将其替换为您的代码。
# 按 双击 Shift 在所有地方搜索类、文件、工具窗口、操作和设置。

SRC_PATH = Path.absolute(Path(__file__)).parent
goodslists = []
goodlist = []


# file_path = str(SRC_PATH / "a/b.ui")
# 不要引相对路径。打包时会有问题，把所有相对路径都改成上面的格式。

def copy_folder(source_folder, destination_folder):
    if not os.path.exists(destination_folder):
        os.mkdir(destination_folder)

    for item in os.listdir(source_folder):
        if os.path.isfile(os.path.join(source_folder, item)):
            shutil.copy(os.path.join(source_folder, item), destination_folder)
        else:
            copy_folder(os.path.join(source_folder, item), os.path.join(destination_folder, item))


class OpeningUI(QDialog):
    def __init__(self):
        super().__init__()
        self.ui = None
        self.init_ui()
        self.folder = None

    def init_ui(self):
        self.ui = uic.loadUi(str(SRC_PATH / "./UI/opening.ui"))

        pushButton = self.ui.pushButton
        pushButton_2 = self.ui.pushButton_2
        pushButton_3 = self.ui.pushButton_3

        pushButton.clicked.connect(lambda: self.importfolder(self.ui.lineEdit_2))
        pushButton_3.clicked.connect(lambda: self.importfolder(self.ui.lineEdit))
        pushButton_2.clicked.connect(lambda: self.loadv3file(self.ui.lineEdit_2, self.ui.lineEdit))

    def importfolder(self, QlineEdit):
        self.folder = QFileDialog.getExistingDirectory(self, "选择文件夹")
        QlineEdit.setText(self.folder)

    def loadv3file(self, QlineEdit1, QlineEdit2):
        GamePath = QlineEdit1.text()
        ModPath = QlineEdit2.text()

        tempmodpath = tempfile.mkdtemp()
        if os.path.exists(tempmodpath + '/common'):
            shutil.rmtree(tempmodpath + '/common')
        copy_folder(GamePath + '/common', tempmodpath + '/common')
        copy_folder(ModPath + '/common', tempmodpath + '/common')
        try:
            BM = BackendManager(GamePath, tempmodpath)
            goods, rowgoods = BM.get_part("goods")
            pm, rowpm = BM.get_part("pm")
            header_labels = ["名称", "价格", "texture", "类型", "威望因数", "消费税",
                             "交易数量", "贸易船乘数", "成瘾机会", "本地商品", "能否交易",
                             "价格固定"]
            # "Name", "Cost", "texture", "category", "prestige_factor", "consumption_tax_cost", "traded_quantity",
            # "convoy_cost_multiplier", "obsession_chance", "本地商品", "无法交易", "价格固定"
            main_ui.ui.tableWidget_goods.setHorizontalHeaderLabels(header_labels)

            k = 0

            for i in goods:
                single_goods, rc = get_goods_detail(BM, i)
                single_goods_str = str(single_goods)
                single_goods_dict = ast.literal_eval(single_goods_str)
                goodslists.append([i, str(single_goods_dict[i][1].get('cost', ["", ""])[1])])
                goodlist.append(str(i))
                main_ui.ui.tableWidget_goods.setItem(k, 0, QTableWidgetItem(i))
                main_ui.ui.tableWidget_goods.setItem(k, 1,
                                                 QTableWidgetItem(str(single_goods_dict[i][1].get('cost', ["", ""])[1])))
                main_ui.ui.tableWidget_goods.setItem(k, 2,
                                                 QTableWidgetItem(str(single_goods_dict[i][1].get('texture', ["", ""])[1])))
                main_ui.ui.tableWidget_goods.setItem(k, 3, QTableWidgetItem(
                    str(single_goods_dict[i][1].get('category', ["", ""])[1])))
                main_ui.ui.tableWidget_goods.setItem(k, 4, QTableWidgetItem(
                    str(single_goods_dict[i][1].get('prestige_factor', ["", ""])[1])))
                main_ui.ui.tableWidget_goods.setItem(k, 5, QTableWidgetItem(
                    str(single_goods_dict[i][1].get('consumption_tax_cost', ["", ""])[1])))
                main_ui.ui.tableWidget_goods.setItem(k, 6, QTableWidgetItem(
                    str(single_goods_dict[i][1].get('traded_quantity', ["", ""])[1])))
                main_ui.ui.tableWidget_goods.setItem(k, 7, QTableWidgetItem(
                    str(single_goods_dict[i][1].get('convoy_cost_multiplier', ["", ""])[1])))
                main_ui.ui.tableWidget_goods.setItem(k, 8, QTableWidgetItem(
                    str(single_goods_dict[i][1].get('obsession_chance', ["", ""])[1])))
                main_ui.ui.tableWidget_goods.setItem(k, 9,
                                                 QTableWidgetItem(str(single_goods_dict[i][1].get('local', ["", ""])[1])))
                main_ui.ui.tableWidget_goods.setItem(k, 10, QTableWidgetItem(
                    str(single_goods_dict[i][1].get('tradeable', ["", ""])[1])))
                main_ui.ui.tableWidget_goods.setItem(k, 11, QTableWidgetItem(
                    str(single_goods_dict[i][1].get('fixed_price', ["", ""])[1])))
                row_count = main_ui.ui.tableWidget_goods.rowCount()
                main_ui.ui.tableWidget_goods.insertRow(row_count)
                k = k + 1
            main_table3.ui.comboBox_3.addItems(goodlist)
            main_table3.ui.comboBox_2.addItems(goodlist)
            header_labels = ["名称", "建筑输入", "建筑输出", "建筑等级规模修正", "建筑非规模修正", '图标路径', '是否默认',
                             '国家劳力规模修正', '国家等级规模修正', '国家非规模修正', '州劳力规模修正', '州等级规模修正',
                             '州非规模修正', '改变PM的修正', '激活需要商品', '解锁法律', '解锁科技', '解锁PM',
                             '解锁全球科技', 'AI权重', '污染值']
            main_ui.ui.tableWidget_pm.setHorizontalHeaderLabels(header_labels)
            k = 0
            for i in pm:
                single_pm, rc = get_pm_detail(BM, i)
                single_pm_str = str(single_pm)
                single_pm_dict = ast.literal_eval(single_pm_str)
                main_ui.ui.tableWidget_pm.setItem(k, 0, QTableWidgetItem(i))

                single_pm_input = single_pm_dict.get('input', {})
                text = ''
                for j in single_pm_input.keys():
                    temptext = '{} = {}\n'.format(j, single_pm_input[j])
                    text = text + temptext
                main_ui.ui.tableWidget_pm.setItem(k, 1, QTableWidgetItem(text))

                single_pm_input = single_pm_dict.get('output', {})
                text = ''
                for j in single_pm_input.keys():
                    temptext = '{} = {}\n'.format(j, single_pm_input[j])
                    text = text + temptext
                main_ui.ui.tableWidget_pm.setItem(k, 2, QTableWidgetItem(text))

                single_pm_input = single_pm_dict.get('level', {})
                if single_pm_input is None:
                    single_pm_input = {}
                text = ''
                for j in single_pm_input.keys():
                    temptext = '{} = {}\n'.format(j, single_pm_input[j][1])
                    text = text + temptext
                main_ui.ui.tableWidget_pm.setItem(k, 3, QTableWidgetItem(text))

                single_pm_input = single_pm_dict.get('unscaled', {})
                if single_pm_input is None:
                    single_pm_input = {}
                text = ''
                for j in single_pm_input.keys():
                    temptext = '{} = {}\n'.format(j, single_pm_input[j][1])
                    text = text + temptext
                main_ui.ui.tableWidget_pm.setItem(k, 4, QTableWidgetItem(text))

                single_pm_input = single_pm_dict.get('texture', "")
                if single_pm_input is None:
                    single_pm_input = ""
                text = str(single_pm_input)
                main_ui.ui.tableWidget_pm.setItem(k, 5, QTableWidgetItem(text))

                single_pm_input = single_pm_dict.get('is_default', "")
                if single_pm_input is None:
                    single_pm_input = ""
                text = str(single_pm_input)
                main_ui.ui.tableWidget_pm.setItem(k, 6, QTableWidgetItem(text))

                single_pm_input = single_pm_dict.get('country_modifiers_workforce', {})
                if single_pm_input is None:
                    single_pm_input = {}
                text = ''
                for j in single_pm_input.keys():
                    temptext = '{} = {}\n'.format(j, single_pm_input[j][1])
                    text = text + temptext
                main_ui.ui.tableWidget_pm.setItem(k, 7, QTableWidgetItem(text))

                single_pm_input = single_pm_dict.get('country_modifiers_level', {})
                if single_pm_input is None:
                    single_pm_input = {}
                text = ''
                for j in single_pm_input.keys():
                    temptext = '{} = {}\n'.format(j, single_pm_input[j][1])
                    text = text + temptext
                main_ui.ui.tableWidget_pm.setItem(k, 8, QTableWidgetItem(text))

                single_pm_input = single_pm_dict.get('country_modifiers_unscaled', {})
                if single_pm_input is None:
                    single_pm_input = {}
                text = ''
                for j in single_pm_input.keys():
                    temptext = '{} = {}\n'.format(j, single_pm_input[j][1])
                    text = text + temptext
                main_ui.ui.tableWidget_pm.setItem(k, 9, QTableWidgetItem(text))

                single_pm_input = single_pm_dict.get('state_modifiers_workforce', {})
                if single_pm_input is None:
                    single_pm_input = {}
                text = ''
                for j in single_pm_input.keys():
                    temptext = '{} = {}\n'.format(j, single_pm_input[j][1])
                    text = text + temptext
                main_ui.ui.tableWidget_pm.setItem(k, 10, QTableWidgetItem(text))

                single_pm_input = single_pm_dict.get('state_modifiers_level', {})
                if single_pm_input is None:
                    single_pm_input = {}
                text = ''
                for j in single_pm_input.keys():
                    temptext = '{} = {}\n'.format(j, single_pm_input[j][1])
                    text = text + temptext
                main_ui.ui.tableWidget_pm.setItem(k, 11, QTableWidgetItem(text))

                single_pm_input = single_pm_dict.get('state_modifiers_unscaled', {})
                if single_pm_input is None:
                    single_pm_input = {}
                text = ''
                for j in single_pm_input.keys():
                    temptext = '{} = {}\n'.format(j, single_pm_input[j][1])
                    text = text + temptext
                main_ui.ui.tableWidget_pm.setItem(k, 12, QTableWidgetItem(text))

                single_pm_input = single_pm_dict.get('timed_modifiers', [])
                if single_pm_input is None:
                    single_pm_input = []
                text = ''
                for j in single_pm_input:
                    temptext = '{}\n'.format(j)
                    text = text + temptext
                main_ui.ui.tableWidget_pm.setItem(k, 13, QTableWidgetItem(text))

                single_pm_input = single_pm_dict.get('required_input_goods', "")
                if single_pm_input is None:
                    single_pm_input = ""
                text = str(single_pm_input)
                main_ui.ui.tableWidget_pm.setItem(k, 14, QTableWidgetItem(text))

                single_pm_input = single_pm_dict.get('unlocking_laws', [])
                if single_pm_input is None:
                    single_pm_input = []
                text = ''
                for j in single_pm_input:
                    temptext = '{}\n'.format(j)
                    text = text + temptext
                main_ui.ui.tableWidget_pm.setItem(k, 15, QTableWidgetItem(text))

                single_pm_input = single_pm_dict.get('unlocking_technologies', [])
                if single_pm_input is None:
                    single_pm_input = []
                text = ''
                for j in single_pm_input:
                    temptext = '{}\n'.format(j)
                    text = text + temptext
                main_ui.ui.tableWidget_pm.setItem(k, 16, QTableWidgetItem(text))

                single_pm_input = single_pm_dict.get('unlocking_production_methods', [])
                if single_pm_input is None:
                    single_pm_input = []
                text = ''
                for j in single_pm_input:
                    temptext = '{}\n'.format(j)
                    text = text + temptext
                main_ui.ui.tableWidget_pm.setItem(k, 17, QTableWidgetItem(text))

                single_pm_input = single_pm_dict.get('unlocking_global_technologies', [])
                if single_pm_input is None:
                    single_pm_input = []
                text = ''
                for j in single_pm_input:
                    temptext = '{}\n'.format(j)
                    text = text + temptext
                main_ui.ui.tableWidget_pm.setItem(k, 18, QTableWidgetItem(text))

                single_pm_input = single_pm_dict.get('ai_weight', "")
                if single_pm_input is None:
                    single_pm_input = ""
                text = str(single_pm_input)
                main_ui.ui.tableWidget_pm.setItem(k, 19, QTableWidgetItem(text))

                single_pm_input = single_pm_dict.get('pollution_generation', "")
                if single_pm_input is None:
                    single_pm_input = ""
                text = str(single_pm_input)
                main_ui.ui.tableWidget_pm.setItem(k, 20, QTableWidgetItem(text))

                row_count = main_ui.ui.tableWidget_pm.rowCount()  # 返回当前行数(尾部)
                main_ui.ui.tableWidget_pm.insertRow(row_count)
                k = k + 1
            shutil.rmtree(tempmodpath + '/common')
            opening.ui.hide()
            main_ui.ui.show()
        except Exception as e:
            print(str(e))
        #  隐藏并显示MAIN界面。


class MainUI(QMainWindow):
    # 幽默主界面

    # 实体化 #
    def __init__(self):
        super().__init__()
        self.ui = None
        self.init_ui()

    # 定义界面 #
    def init_ui(self):
        self.ui = uic.loadUi(str(SRC_PATH / "./UI/main.ui"))

        # 调取UI控件
        pushButton_load_goods = self.ui.pushButton_load_goods
        pushButton_load_pm = self.ui.pushButton_load_pm
        pushButton_load_pmg = self.ui.pushButton_load_pmg
        pushButton_load_bg = self.ui.pushButton_load_bg
        pushButton_pm_tools = self.ui.pushButton_pm_tools
        pushButton_tools_export_goods = self.ui.pushButton_tools_export_goods
        pushButton_tools_import_goods = self.ui.pushButton_tools_import_goods
        pushButton_tools_about = self.ui.pushButton_tools_about
        pushButton_goods_export = self.ui.pushButton_goods_export
        pushButton_goods_add = self.ui.pushButton_goods_add
        pushButton_goods_del = self.ui.pushButton_goods_del
        tableWidget_goods = self.ui.tableWidget_goods
        tableWidget_pm = self.ui.tableWidget_pm
        pushButton_pm_export = self.ui.pushButton_pm_export
        pushButton_pm_add = self.ui.pushButton_pm_add
        pushButton_pm_del = self.ui.pushButton_pm_del
        pushButton_pm_save = self.ui.pushButton_pm_save

        # 定义信号，连接槽函数
        pushButton_load_goods.clicked.connect(lambda: self.changepage_goods(self.ui.stackedWidget))
        pushButton_load_pm.clicked.connect(lambda: self.changepage_pm(self.ui.stackedWidget))
        pushButton_load_pmg.clicked.connect(lambda: self.changepage_pmg(self.ui.stackedWidget))
        pushButton_load_bg.clicked.connect(lambda: self.changepage_bg(self.ui.stackedWidget))
        pushButton_pm_tools.clicked.connect(self.pm_tools)
        pushButton_tools_export_goods.clicked.connect(self.tools_export_goods)
        pushButton_tools_import_goods.clicked.connect(self.tools_import_goods)
        pushButton_tools_about.clicked.connect(self.tools_about)
        pushButton_goods_del.clicked.connect(self.goodsdelete)
        pushButton_goods_add.clicked.connect(self.goodsadd)
        pushButton_goods_export.clicked.connect(self.goodsexport)
        pushButton_pm_del.clicked.connect(self.pmdelete)
        pushButton_pm_add.clicked.connect(self.pmadd)
        pushButton_pm_export.clicked.connect(self.pmexport)
        pushButton_pm_save.clicked.connect(self.pmsave)
        tableWidget_pm.itemSelectionChanged.connect(self.pm_itemchange)

    # 定义槽函数 #
    def changepage_goods(self, QstackedWidget):
        QstackedWidget.setCurrentIndex(0)
        main_ui.ui.pushButton_load_goods.setChecked(True)
        main_ui.ui.pushButton_load_pm.setChecked(False)
        main_ui.ui.pushButton_load_pmg.setChecked(False)
        main_ui.ui.pushButton_load_bg.setChecked(False)

    def changepage_pm(self, QstackedWidget):
        QstackedWidget.setCurrentIndex(1)
        main_ui.ui.pushButton_load_goods.setChecked(False)
        main_ui.ui.pushButton_load_pm.setChecked(True)
        main_ui.ui.pushButton_load_pmg.setChecked(False)
        main_ui.ui.pushButton_load_bg.setChecked(False)

    def changepage_pmg(self, QstackedWidget):
        QstackedWidget.setCurrentIndex(2)
        main_ui.ui.pushButton_load_goods.setChecked(False)
        main_ui.ui.pushButton_load_pm.setChecked(False)
        main_ui.ui.pushButton_load_pmg.setChecked(True)
        main_ui.ui.pushButton_load_bg.setChecked(False)
    def changepage_bg(self, QstackedWidget):
        QstackedWidget.setCurrentIndex(3)
        main_ui.ui.pushButton_load_goods.setChecked(False)
        main_ui.ui.pushButton_load_pm.setChecked(False)
        main_ui.ui.pushButton_load_pmg.setChecked(False)
        main_ui.ui.pushButton_load_bg.setChecked(True)

    def goodsdelete(self):
        row = main_ui.ui.tableWidget_goods.currentRow()
        main_ui.ui.tableWidget_goods.removeRow(row)

    def goodsadd(self):
        row_count = main_ui.ui.tableWidget_goods.rowCount()
        main_ui.ui.tableWidget_goods.insertRow(row_count)

    def goodsexport(self):
        ModPath = opening.ui.lineEdit.text()
        shutil.rmtree(ModPath + '/common/goods')
        os.mkdir(ModPath + '/common/goods')
        with open(ModPath + '/common/goods/00_goods.txt', 'w+') as f:
            f.close()
        with open(ModPath + '/common/goods/goods.txt', 'w+') as f:
            row_count = main_ui.ui.tableWidget_goods.rowCount()
            text = ''
            temp = ''
            # "Name", "Cost", "texture", "category", "prestige_factor", "consumption_tax_cost", "traded_quantity",
            # "convoy_cost_multiplier", "obsession_chance", "本地商品", "无法交易", "价格固定"
            for i in range(0, row_count - 1):
                if main_ui.ui.tableWidget_goods.item(i, 0).text() != '':
                    temp = temp + '{} = {{\n'.format(main_ui.ui.tableWidget_goods.item(i, 0).text())
                    if main_ui.ui.tableWidget_goods.item(i, 1).text() != '':
                        temp = temp + '\tcost = {}\n'.format(main_ui.ui.tableWidget_goods.item(i, 1).text())
                    if main_ui.ui.tableWidget_goods.item(i, 2).text() != '':
                        temp = temp + '\ttexture = {}\n'.format(main_ui.ui.tableWidget_goods.item(i, 2).text())
                    if main_ui.ui.tableWidget_goods.item(i, 3).text() != '':
                        temp = temp + '\tcategory = {}\n'.format(main_ui.ui.tableWidget_goods.item(i, 3).text())
                    if main_ui.ui.tableWidget_goods.item(i, 4).text() != '':
                        temp = temp + '\tprestige_factor = {}\n'.format(main_ui.ui.tableWidget_goods.item(i, 4).text())
                    if main_ui.ui.tableWidget_goods.item(i, 5).text() != '':
                        temp = temp + '\tconsumption_tax_cost = {}\n'.format(main_ui.ui.tableWidget_goods.item(i, 5).text())
                    if main_ui.ui.tableWidget_goods.item(i, 6).text() != '':
                        temp = temp + '\ttraded_quantity = {}\n'.format(main_ui.ui.tableWidget_goods.item(i, 6).text())
                    if main_ui.ui.tableWidget_goods.item(i, 7).text() != '':
                        temp = temp + '\tconvoy_cost_multiplier = {}\n'.format(
                            main_ui.ui.tableWidget_goods.item(i, 7).text())
                    if main_ui.ui.tableWidget_goods.item(i, 8).text() != '':
                        temp = temp + '\tobsession_chance = {}\n'.format(main_ui.ui.tableWidget_goods.item(i, 8).text())
                    if main_ui.ui.tableWidget_goods.item(i, 9).text() != '':
                        temp = temp + '\tlocal = {}\n'.format(main_ui.ui.tableWidget_goods.item(i, 9).text())
                    if main_ui.ui.tableWidget_goods.item(i, 10).text() != '':
                        temp = temp + '\ttradeable = {}\n'.format(main_ui.ui.tableWidget_goods.item(i, 10).text())
                    if main_ui.ui.tableWidget_goods.item(i, 11).text() != '':
                        temp = temp + '\tfixed_price = {}\n'.format(main_ui.ui.tableWidget_goods.item(i, 11).text())
                    temp = temp + '}\n'
                text = text + temp
                temp = ''
            f.write(text)
            f.close()

    def pmdelete(self):
        row = main_ui.ui.tableWidget_pm.currentRow()
        main_ui.ui.tableWidget_pm.removeRow(row)

    def pmadd(self):
        row_count = main_ui.ui.tableWidget_pm.rowCount()
        main_ui.ui.tableWidget_pm.insertRow(row_count)

    def pmsave(self):
        row = main_ui.ui.tableWidget_pm.currentRow()
        column = main_ui.ui.tableWidget_pm.currentColumn()
        main_ui.ui.tableWidget_pm.setItem(row, column, QTableWidgetItem(main_ui.ui.textEdit_pm.toPlainText()))

    def pmexport(self):
        ModPath = opening.ui.lineEdit.text()
        shutil.rmtree(ModPath + '/common/production_methods')
        os.mkdir(ModPath + '/common/production_methods')
        with open(ModPath + '/common/production_methods/00_dummy.txt', 'w+') as f:
            f.close()
        with open(ModPath + '/common/production_methods/01_industry.txt', 'w+') as f:
            f.close()
        with open(ModPath + '/common/production_methods/02_agro.txt', 'w+') as f:
            f.close()
        with open(ModPath + '/common/production_methods/03_mines.txt', 'w+') as f:
            f.close()
        with open(ModPath + '/common/production_methods/04_plantations.txt', 'w+') as f:
            f.close()
        with open(ModPath + '/common/production_methods/05_military.txt', 'w+') as f:
            f.close()
        with open(ModPath + '/common/production_methods/06_urban_center.txt', 'w+') as f:
            f.close()
        with open(ModPath + '/common/production_methods/07_government.txt', 'w+') as f:
            f.close()
        with open(ModPath + '/common/production_methods/08_monuments.txt', 'w+') as f:
            f.close()
        with open(ModPath + '/common/production_methods/09_misc_resource.txt', 'w+') as f:
            f.close()
        with open(ModPath + '/common/production_methods/10_canals.txt', 'w+') as f:
            f.close()
        with open(ModPath + '/common/production_methods/11_private_infrastructure.txt', 'w+') as f:
            f.close()
        with open(ModPath + '/common/production_methods/12_subsistence.txt', 'w+') as f:
            f.close()
        with open(ModPath + '/common/production_methods/13_construction.txt', 'w+') as f:
            f.close()
        with open(ModPath + '/common/production_methods/pm.txt', 'w+') as f:
            # ["0名称", "1建筑输入", "2建筑输出", "3建筑等级规模修正", "4建筑非规模修正", '5图标路径', '6是否默认',
            # '7国家劳力规模修正', '8国家等级规模修正', '9国家非规模修正', '10州劳力规模修正', '11州等级规模修正',
            # '12州非规模修正', '13改变PM的修正', '14激活需要商品', '15解锁法律', '16解锁科技', '17解锁PM',
            # '18解锁全球科技', '19AI权重', '20污染值']
            row_count = main_ui.ui.tableWidget_pm.rowCount()
            text = ''
            temp = ''
            for i in range(0, row_count - 1):
                if main_ui.ui.tableWidget_pm.item(i, 0).text() != '':
                    temp = temp + '{} = {{\n'.format(main_ui.ui.tableWidget_pm.item(i, 0).text())
                    if main_ui.ui.tableWidget_pm.item(i, 5).text() != '':
                        temp = temp + '\ttexture = {}\n'.format(main_ui.ui.tableWidget_pm.item(i, 5).text())
                    if main_ui.ui.tableWidget_pm.item(i, 6).text() != '':
                        temp = temp + '\tis_default = {}\n'.format(main_ui.ui.tableWidget_pm.item(i, 6).text())
                    if main_ui.ui.tableWidget_pm.item(i, 14).text() != '':
                        temp = temp + '\trequired_input_goods = {}\n'.format(main_ui.ui.tableWidget_pm.item(i, 14).text())
                    if main_ui.ui.tableWidget_pm.item(i, 19).text() != '':
                        temp = temp + '\tai_weight = {}\n'.format(main_ui.ui.tableWidget_pm.item(i, 19).text())
                    if main_ui.ui.tableWidget_pm.item(i, 20).text() != '':
                        temp = temp + '\tpollution_generation = {}\n'.format(main_ui.ui.tableWidget_pm.item(i, 20).text())
                    if main_ui.ui.tableWidget_pm.item(i, 15).text() != '':
                        temp = temp + '\tunlocking_laws = {\n'+main_ui.ui.tableWidget_pm.item(i, 15).text()+'\n\t}\n'
                    if main_ui.ui.tableWidget_pm.item(i, 16).text() != '':
                        temp = temp + '\tunlocking_technologies = {\n'+main_ui.ui.tableWidget_pm.item(i, 16).text()+'\n\t}\n'
                    if main_ui.ui.tableWidget_pm.item(i, 17).text() != '':
                        temp = temp + '\tunlocking_production_methods = {\n'+main_ui.ui.tableWidget_pm.item(i, 17).text()+'\n\t}\n'
                    if main_ui.ui.tableWidget_pm.item(i, 18).text() != '':
                        temp = temp + '\tunlocking_global_technologies = {\n'+main_ui.ui.tableWidget_pm.item(i, 18).text()+'\n\t}\n'
                    if main_ui.ui.tableWidget_pm.item(i, 13).text() != '':
                        temp = temp + '\ttimed_modifiers = {\n'+main_ui.ui.tableWidget_pm.item(i, 13).text()+'\n\t}\n'
                    if main_ui.ui.tableWidget_pm.item(i, 7).text() != '' or main_ui.ui.tableWidget_pm.item(i, 8).text() != '' or main_ui.ui.tableWidget_pm.item(i, 9).text() != '':
                        temp = temp + '\tcountry_modifiers = {\n'
                        if main_ui.ui.tableWidget_pm.item(i, 7).text() != '':
                            temp = temp + '\tworkforce_scaled = {\n'+main_ui.ui.tableWidget_pm.item(i, 7).text()+'\n\t}\n'
                        if main_ui.ui.tableWidget_pm.item(i, 8).text() != '':
                            temp = temp + '\tlevel_scaled = {\n'+main_ui.ui.tableWidget_pm.item(i, 8).text()+'\n\t}\n'
                        if main_ui.ui.tableWidget_pm.item(i, 9).text() != '':
                            temp = temp + '\tunscaled = {\n'+main_ui.ui.tableWidget_pm.item(i, 9).text()+'\n\t}\n'
                        temp = temp + '\t}\n'
                    if main_ui.ui.tableWidget_pm.item(i, 10).text() != '' or main_ui.ui.tableWidget_pm.item(i, 11).text() != '' or main_ui.ui.tableWidget_pm.item(i, 12).text() != '':
                        temp = temp + '\tstate_modifiers = {\n'
                        if main_ui.ui.tableWidget_pm.item(i, 10).text() != '':
                            temp = temp + '\tworkforce_scaled = {\n'+main_ui.ui.tableWidget_pm.item(i, 10).text()+'\n\t}\n'
                        if main_ui.ui.tableWidget_pm.item(i, 11).text() != '':
                            temp = temp + '\tlevel_scaled = {\n'+main_ui.ui.tableWidget_pm.item(i, 11).text()+'\n\t}\n'
                        if main_ui.ui.tableWidget_pm.item(i, 12).text() != '':
                            temp = temp + '\tunscaled = {\n'+main_ui.ui.tableWidget_pm.item(i, 12).text()+'\n\t}\n'
                        temp = temp + '\t}\n'
                    if main_ui.ui.tableWidget_pm.item(i, 1).text() != '' or main_ui.ui.tableWidget_pm.item(i, 2).text() != '' or main_ui.ui.tableWidget_pm.item(i, 3).text() != '' or main_ui.ui.tableWidget_pm.item(i, 4).text() != '':
                        temp = temp + '\tbuilding_modifiers = {\n'
                        if main_ui.ui.tableWidget_pm.item(i, 1).text() != '' or main_ui.ui.tableWidget_pm.item(i, 2).text() != '':
                            temp = temp + '\tworkforce_scaled = {\n'
                            if main_ui.ui.tableWidget_pm.item(i, 1).text() != '':
                                strlist = main_ui.ui.tableWidget_pm.item(i, 1).text().splitlines()
                                for j in strlist:
                                    if j != '':
                                        j = j.replace(" ","")
                                        strlist2 = j.split("=")
                                        temp = temp + '\t\tgoods_input_' + strlist2[0] + '_add = ' + strlist2[1] + '\n'
                                temp = temp + ''
                            if main_ui.ui.tableWidget_pm.item(i, 2).text() != '':
                                strlist = main_ui.ui.tableWidget_pm.item(i, 2).text().splitlines()
                                for j in strlist:
                                    if j != '':
                                        j = j.replace(" ","")
                                        strlist2 = j.split("=")
                                        temp = temp + '\t\tgoods_output_' + strlist2[0] + '_add = ' + strlist2[1] + '\n'
                                temp = temp + ''
                            temp = temp + '\t}\n'
                        if main_ui.ui.tableWidget_pm.item(i, 3).text() != '':
                            temp = temp + '\tlevel_scaled = {\n'+main_ui.ui.tableWidget_pm.item(i, 3).text()+'\n\t}\n'
                        if main_ui.ui.tableWidget_pm.item(i, 4).text() != '':
                            temp = temp + '\tunscaled = {\n'+main_ui.ui.tableWidget_pm.item(i, 4).text()+'\n\t}\n'
                        temp = temp + '\t}\n'
                    temp = temp + '}\n'
                text = text + temp
                temp = ''
            f.write(text)
            f.close()

    def pm_itemchange(self):
        text = str(main_ui.ui.tableWidget_pm.currentItem().text())
        main_ui.ui.textEdit_pm.setText(text)

    def tools_export_goods(self):
        # 幽默展示和隐藏
        main_table1.ui.show()
        main_table2.ui.hide()
        main_table3.ui.hide()
        # ...

    def tools_import_goods(self):
        main_table1.ui.hide()
        main_table2.ui.show()
        main_table3.ui.hide()

    def pm_tools(self):
        main_table1.ui.hide()
        main_table2.ui.hide()
        main_table3.ui.show()

    def tools_about(self):
        msgbox = QMessageBox(QMessageBox.Information, "关于", "V3编辑器\n由小草青青、VK8040、恋恋的钢盔共同制作",
                             QMessageBox.Ok, self)
        msgbox.show()


class MainTable_export_goods(QWidget):
    def __init__(self):
        super().__init__()
        self.ui = None
        self.folder = None
        self.init_ui()

    def init_ui(self):
        self.ui = uic.loadUi(str(SRC_PATH / "./UI/export_goods.ui"))
        self.ui.pushButton.clicked.connect(self.button_export_path)
        self.ui.pushButton_2.clicked.connect(self.button_download)

    def button_export_path(self):
        self.folder = QFileDialog.getExistingDirectory(self, "选择文件夹")
        self.ui.lineEdit.setText(self.folder)

    def button_download(self):
        # 源文件路径
        self.folder = self.ui.lineEdit.text()
        source_file = str(SRC_PATH / "./UI/商品模版.xlsx")

        # 目标文件路径（包括新的文件名）
        target_file = str(self.folder) + "/商品模版.xlsx"

        try:
            # 复制文件
            shutil.copy(source_file, target_file)
            msgbox = QMessageBox(QMessageBox.Information, "成功", "成功导出！",
                                 QMessageBox.Ok, self)
            msgbox.show()
        except FileNotFoundError:
            msgbox = QMessageBox(QMessageBox.Information, "错误", "没有找到路径",
                                 QMessageBox.Ok, self)
            msgbox.show()
        except PermissionError:
            msgbox = QMessageBox(QMessageBox.Information, "错误", "没有足够权限",
                                 QMessageBox.Ok, self)
            msgbox.show()
        except Exception as e:
            msgbox = QMessageBox(QMessageBox.Information, "错误", "发生错误：" + str(e),
                                 QMessageBox.Ok, self)
            msgbox.show()


class MainTable_import_goods(QWidget):
    def __init__(self):
        super().__init__()
        self.ui = None
        self.folder = None
        self.folder2 = None
        self.init_ui()

    def init_ui(self):
        self.ui = uic.loadUi(str(SRC_PATH / "./UI/import_goods.ui"))
        self.ui.pushButton.clicked.connect(self.pushbutton_clicked)
        self.ui.pushButton_2.clicked.connect(self.pushbutton_2_clicked)
        self.ui.pushButton_3.clicked.connect(self.pushbutton_3_clicked)

    def pushbutton_clicked(self):
        self.folder = QFileDialog.getOpenFileName(self, "打开文件", "./")
        self.ui.lineEdit.setText(self.folder[0])

    def pushbutton_2_clicked(self):
        self.folder2 = QFileDialog.getSaveFileName(self, "文件保存为", "./", "txt File (*.txt)")
        self.ui.lineEdit_2.setText(self.folder2[0])

    def pushbutton_3_clicked(self):
        fd = self.ui.lineEdit.text()
        fd2 = self.ui.lineEdit_2.text()
        wb = op.load_workbook(fd)
        ws = wb.worksheets[0]
        text = ''
        max_row = ws.max_row
        for i in range(2, max_row + 1):
            temp = '{} = {{\n\ttexture = {}\n\tcost = {}\n\tcategory = {}\n\tprestige_factor = {}\n' \
                .format(ws['C' + str(i)].value, ws['E' + str(i)].value, ws['D' + str(i)].value, ws['F' + str(i)].value,
                        ws['G' + str(i)].value)
            if ws['H' + str(i)].value is None:
                temp = temp
            else:
                temp = temp + '\tconsumption_tax_cost = {}\n' \
                    .format(ws['H' + str(i)].value)
            if ws['I' + str(i)].value is None:
                temp = temp
            else:
                temp = temp + '\ttraded_quantity = {}\n' \
                    .format(ws['I' + str(i)].value)
            if ws['J' + str(i)].value is None:
                temp = temp
            else:
                temp = temp + '\tconvoy_cost_multiplier = {}\n' \
                    .format(ws['J' + str(i)].value)
            if ws['K' + str(i)].value is None:
                temp = temp
            else:
                temp = temp + '\tobsession_chance = {}\n' \
                    .format(ws['K' + str(i)].value)
            if ws['L' + str(i)].value is None:
                temp = temp
            else:
                temp = temp + '\tlocal = {}\n' \
                    .format(ws['L' + str(i)].value)
            temp = temp + '}\n'
            text = text + temp
        with open(fd2, 'w', encoding='utf_8_sig') as fp:
            fp.write(text)
            fp.close()


class MainTable_PM(QDialog):
    def __init__(self):
        super().__init__(main_ui)
        self.ui = None
        self.init_ui()

    def init_ui(self):
        self.ui = uic.loadUi(str(SRC_PATH / "./UI/pm.ui"))
        pushButton_inadd = self.ui.pushButton_inadd
        pushButton_indel = self.ui.pushButton_indel
        pushButton_outadd = self.ui.pushButton_outadd
        pushButton_outdel = self.ui.pushButton_outdel
        pushButton_insave = self.ui.pushButton_insave
        pushButton_outsave = self.ui.pushButton_outsave
        pushButton = self.ui.pushButton
        header_labels = ["名称", "数量", "价格", "总价", "占比"]
        self.ui.tableWidget.setHorizontalHeaderLabels(header_labels)
        self.ui.tableWidget_goods.setHorizontalHeaderLabels(header_labels)

        pushButton_inadd.clicked.connect(self.in_add)
        pushButton_indel.clicked.connect(self.in_del)
        pushButton_outadd.clicked.connect(self.out_add)
        pushButton_outdel.clicked.connect(self.out_del)
        pushButton_insave.clicked.connect(self.save)
        pushButton_outsave.clicked.connect(self.save)
        pushButton.clicked.connect(self.copy)

    def in_add(self):
        in_int = self.ui.comboBox_3.currentIndex()
        row_count = self.ui.tableWidget.rowCount()
        self.ui.tableWidget.setItem(row_count - 1, 0, QTableWidgetItem(str(goodslists[in_int][0])))
        self.ui.tableWidget.setItem(row_count - 1, 1, QTableWidgetItem(self.ui.lineEdit_4.text()))
        self.ui.tableWidget.setItem(row_count - 1, 2, QTableWidgetItem(str(goodslists[in_int][1])))
        self.ui.tableWidget.insertRow(row_count)
        print("A")
        self.recount()

    def in_del(self):
        row = self.ui.tableWidget.currentRow()
        self.ui.tableWidget.removeRow(row)
        self.recount()

    def out_add(self):
        out_int = self.ui.comboBox_2.currentIndex()
        row_count = self.ui.tableWidget_goods.rowCount()
        self.ui.tableWidget_goods.setItem(row_count - 1, 0, QTableWidgetItem(str(goodslists[out_int][0])))
        self.ui.tableWidget_goods.setItem(row_count - 1, 1, QTableWidgetItem(self.ui.lineEdit_3.text()))
        self.ui.tableWidget_goods.setItem(row_count - 1, 2, QTableWidgetItem(str(goodslists[out_int][1])))
        self.ui.tableWidget_goods.insertRow(row_count)
        self.recount()

    def out_del(self):
        row = self.ui.tableWidget_goods.currentRow()
        self.ui.tableWidget_goods.removeRow(row)
        self.recount()

    def save(self):
        self.recount()

    def recount(self):
        row_in = self.ui.tableWidget.rowCount()

        total_in = 0
        for i in range(0, row_in - 1):
            if self.ui.tableWidget.item(i, 1).text() != "" and self.ui.tableWidget.item(i, 2).text() != "":

                try:
                    total_in_good = float(self.ui.tableWidget.item(i, 1).text()) * float(
                        self.ui.tableWidget.item(i, 2).text())
                    self.ui.tableWidget.setItem(i, 3, QTableWidgetItem(str(total_in_good)))

                except Exception as e:
                    total_in_good = 0
                    self.ui.tableWidget.setItem(i, 3, QTableWidgetItem("0"))
            else:
                total_in_good = 0
                self.ui.tableWidget.setItem(i, 3, QTableWidgetItem("0"))
            total_in = total_in + total_in_good
        self.ui.lineEdit.setText(str(total_in))
        if total_in != 0:
            for i in range(0, row_in - 1):
                in_per = float(self.ui.tableWidget.item(i, 3).text()) / total_in
                self.ui.tableWidget.setItem(i, 4, QTableWidgetItem(str(in_per)))

        row_out = self.ui.tableWidget_goods.rowCount()

        total_out = 0
        for i in range(0, row_out - 1):
            if self.ui.tableWidget_goods.item(i, 1).text() != "" and self.ui.tableWidget_goods.item(i, 2).text() != "":
                try:
                    total_out_good = float(self.ui.tableWidget_goods.item(i, 1).text()) * float(
                        self.ui.tableWidget_goods.item(i, 2).text())
                    self.ui.tableWidget_goods.setItem(i, 3, QTableWidgetItem(str(total_out_good)))
                except Exception as e:
                    total_out_good = 0
                    self.ui.tableWidget_goods.setItem(i, 3, QTableWidgetItem("0"))
            else:
                total_out_good = 0
                self.ui.tableWidget_goods.setItem(i, 3, QTableWidgetItem("0"))
            total_out = total_out + total_out_good
        self.ui.lineEdit_6.setText(str(total_out))
        if total_out != 0:
            for i in range(0, row_out - 1):
                out_per = float(self.ui.tableWidget_goods.item(i, 3).text()) / total_out
                self.ui.tableWidget_goods.setItem(i, 4, QTableWidgetItem(str(out_per)))
        total_l = total_out - total_in
        self.ui.lineEdit_8.setText(str(total_l))
        try:
            lirunlv = (total_out - total_in) / total_in
        except ZeroDivisionError as e:
            lirunlv = 0
        lirunlv = lirunlv * 100
        self.ui.lineEdit_9.setText(str(lirunlv))

    def copy(self):
        text = ''
        text = text + 'PM_' + self.ui.lineEdit_2.text() + ' = {\n\ttexture = "gfx/interface/icons/production_method_icons/"' + self.ui.lineEdit_2.text() + '.dds\n\n\t'
        text = text + ' building_modifiers = {\n\t\tworkforce_scaled = {\n\t\t\t#input goods\n\t\t\t'
        row_in = self.ui.tableWidget.rowCount()
        row_out = self.ui.tableWidget_goods.rowCount()
        for i in range(0, row_in - 1):
            text = text + 'goods_input_' + self.ui.tableWidget.item(i, 0).text() + '_add = ' + self.ui.tableWidget.item(
                i, 1).text() + '\n\t\t\t'
        text = text + '#output goods\n\t\t\t'
        for i in range(0, row_out - 1):
            text = text + 'goods_output_' + self.ui.tableWidget_goods.item(i,
                                                                       0).text() + '_add = ' + self.ui.tableWidget_goods.item(
                i, 1).text() + '\n\t\t\t'
        text = text + '}\n\t\t}\n\t}\n'
        self.ui.TextEdit.setText(text)
        pyperclip.copy(text)


def print_hi(name):
    # 在下面的代码行中使用断点来调试脚本。
    print(f'Hi, {name}')  # 按 Ctrl+F8 切换断点。


# 按装订区域中的绿色按钮以运行脚本。
if __name__ == '__main__':
    app = QApplication(sys.argv)

    opening = OpeningUI()
    main_ui = MainUI()
    main_table1 = MainTable_export_goods()
    main_table2 = MainTable_import_goods()
    main_table3 = MainTable_PM()
    opening.ui.show()

    app.exec()
